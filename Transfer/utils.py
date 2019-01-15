import os
import gzip
import datetime

from django.db import IntegrityError, transaction
from fuzzywuzzy import fuzz
from openpyxl import load_workbook

from .transfer_settings import *
from .models import ComponentInformation, TubeInformation, CoreData, ExecutionStats
from .constants import PROJECT_STORAGE

class SubmissionExcelParser(object):
    """Parses customer submission sheet

    First searches sheet for column values then sequentially
    reads sample information from each row. Appropriate
    database objects are created for every sample.

    Error handling done at view level - see Transfer.views
    """


    def __init__(self, sub_sheet_path):
        """Finds Project ID and submission type. """

        self.sub_sheet_path = sub_sheet_path
        self.wb = load_workbook(filename=sub_sheet_path)
        self.general_info_ws = self.wb['General Information']

        for row in range(1, general_sheet_length):
            for column in range(1, general_sheet_width):

                gen_val = cell_value(
                    column,
                    row,
                    self.general_info_ws
                    )

                if gen_val == STR_SAMPLE_TYPE:
                    self.submission_type = cell_value(
                        column,
                        row + 1,
                        worksheet=self.general_info_ws)
                if gen_val == STR_PROJECT_ID:
                    self.project_id_from_sheet = cell_value(
                        column,
                        row + 1,
                        worksheet=self.general_info_ws)


    def find_columns(self):
        """Loads correct worksheet and pulls x, y column coordinates. """

        self.submission_detail_ws = self.wb[self.submission_type]
        self.header_coordinates = {}

        for row in range(1, max_sheet_length):
            for column in range(1, max_sheet_width):

                for header in column_headers:
                    if cell_value(
                        column,
                        row,
                        self.submission_detail_ws
                        ) == header:
                        self.header_coordinates[header] = (column, row)

    @transaction.atomic
    def parse_individual_library_submission(self):
        """Details from "Indexing Information" for individual samples

        'sample_coord[1]' refers to the row that the 'Sample ID' cell is located on
        + 2 is added to the sample row coordinate in order to skip the example row provided
        in the sample submission sheet.
        The range goes to + 100 arbitrarily as the loop will stop if it encounters
        empty cells in the sample column.
        """

        def indiv_cell_val(header):
            return cell_value(
                self.header_coordinates[header][0],
                row,
                self.submission_detail_ws)

        data_start_row = self.header_coordinates[INDIV_ANCHOR][1] + 2
        for row in range(data_start_row, max_sheet_length):

            tube_field_dict = {}

            tube_field_dict['project_id']    = self.project_id_from_sheet
            tube_field_dict['tube_id']       = indiv_cell_val(STR_TUBE_ID)
            tube_field_dict['volume']        = indiv_cell_val(STR_VOLUME)
            tube_field_dict['concentration'] = indiv_cell_val(STR_CONCENTRATION)

            data = TubeInformation.objects.create(**tube_field_dict)

            component_field_dict = {}

            component_field_dict['project_id']        = self.project_id_from_sheet
            component_field_dict['sample_id']         = indiv_cell_val(STR_LIBRARY_ID)
            component_field_dict['i7_index_name']     = indiv_cell_val(STR_I7_INDEX_NAME)
            component_field_dict['i7_index_sequence'] = indiv_cell_val(STR_I7_INDEX_SEQ)
            component_field_dict['i5_index_name']     = indiv_cell_val(STR_I5_INDEX_NAME)
            component_field_dict['i5_index_sequence'] = indiv_cell_val(STR_I5_INDEX_SEQ)

            data = ComponentInformation.objects.create(**component_field_dict)

    @transaction.atomic
    def parse_pool_submission(self):
        """Details from "Pool Submission" sheet

         'sample_coord[1]' refers to the row that the 'Sample ID' cell is located on
         + 2 is added to the sample row coordinate in order to skip the example row provided
         in the sample submission sheet.
         The range goes to + 100 arbitrarily as the loop will stop if it encounters
         empty cells in the sample column.
        """

        for row in range(self.tube_coord[1] + 2, self.tube_coord[1]+20):

            if self.value_in_column(row, self.tube_coord):

                tube_field_dict = {}

                tube_field_dict['project_id']               = self.project_id_from_sheet
                tube_field_dict['tube_id']             = self.value_in_column(row, self.tube_coord)
                tube_field_dict['pool_id']             = self.value_in_column(row, self.pools_coord)
                tube_field_dict['volume']              = self.value_in_column(row, self.volume_coord)
                tube_field_dict['concentration']       = self.value_in_column(row, self.conc_coord)
                tube_field_dict['total_amount']        = self.value_in_column(row, self.amount_coord)
                tube_field_dict['quantitation_method'] = self.value_in_column(row, self.quant_coord)
                tube_field_dict['buffer']              = self.value_in_column(row, self.buff_coord)
                tube_field_dict['organism']            = self.value_in_column(row, self.org_coord)

                data = TubeInformation.objects.create(**tube_field_dict)

        for row in range(self.pool_coord[1] + 2, self.pool_coord[1]+500):
            if self.value_in_column(row, self.sample_coord):

                component_field_dict = {}

                component_field_dict['project_id']             = self.project_id_from_sheet
                component_field_dict['pool_id']           = self.value_in_column(row, self.pool_coord)
                component_field_dict['sample_id']         = self.value_in_column(row, self.sample_coord)
                component_field_dict['i7_index_name']     = self.value_in_column(row, self.i7_name_coord)
                component_field_dict['i7_index_sequence'] = self.value_in_column(row, self.i7_seq_coord)
                component_field_dict['i5_index_name']     = self.value_in_column(row, self.i5_name_coord)
                component_field_dict['i5_index_sequence'] = self.value_in_column(row, self.i5_seq_coord)

                data = ComponentInformation.objects.create(**component_field_dict)

class FastQParser(object):
    """Function to parse the file names and content of Fastq files.

    Create database objects out of the information.
    Example file to be parsed is: 17127FL-27-02-bkbk12-A1_S193_L005_R1_001.fastq.gz
    """
    def __init__(self, fastq_directory, project_id):
        self.fastq_directory = fastq_directory
        self.project_id = project_id

    @staticmethod
    def index_match(string):
        """Raise exception if out of place character appears in sequence"""
        allowed_characters = ['N','A','C','G','T']
        for base in string:
            if base not in allowed_characters:
                raise Exception(f'Non-base character in index: {base}')

    def parse_illumina_fastq_content(fastq_file_obj, core_field_dict):
        """Parser for the contents of any modern Illumina fastQ file
        'top_seq_identifier' refers to the first sequence identifier, of the form:
        @E00558:209:HMKJCCCXY:5:1101:10044:1379 1:N:0:NCTCGCTA+NTAGAGAG
        """

        line_list = []

        # Creates a sub-list of run identifier lines
        for line in fastq_file_obj:
           if line.startswith('@') and len(line_list) < 2:
                line_list.append(line.rstrip('\n '))

        i7_indexes = []
        i5_indexes = []
        flowcell_ids = []

        for line in line_list:

            # Split sequencing identifier into a list of components
            seq_id_list = line.split(":")

            flowcell_ids.append(seq_id_list[2].rstrip('\r\n '))

            # Check if one or two barcodes are present
            if "+" in seq_id_list[-1]:

                i7_index_seq = seq_id_list[-1].split("+")[0].rstrip('\r\n ')
                try:
                    index_match(i7_index_seq)
                except:
                    raise
                i7_indexes.append(i7_index_seq)

                i5_index_seq = seq_id_list[-1].split("+")[1].rstrip('\r\n ')
                try:
                    index_match(i5_index_seq)
                except:
                    raise
                i5_indexes.append(i5_index_seq)

            else:
                i7_index_seq = seq_id_list[-1].rstrip('\r\n ')
                try:
                    index_match(i7_index_seq)
                except:
                    raise
                i7_indexes.append(i7_index_seq)

        # Check if FastQ is demultiplexed
        if i7_indexes[0] == i7_indexes[1]:
            core_field_dict['i7_index_sequence'] = i7_indexes[0]
        else:
            raise Exception('I7 indexes not demultiplexed.')

        if i5_indexes[0] is not None:
            if i5_indexes[0] == i5_indexes[1]:
                core_field_dict['i5_index_sequence'] = i5_indexes[0]
            else:
                print ('WARNING: I5 indexes not demultiplexed.')


        # Check if flowcell ID's are consistent
        if flowcell_ids[0] == flowcell_ids[1]:
            core_field_dict['flowcell_id'] = flowcell_ids[0]
        else:
            raise Exception('Flowcell IDs are not consistent in this FastQ')

        return core_field_dict

    @transaction.atomic
    def parse_fastq_filenames(self):

        numof_files_parsed = 0

        for fastq_file in os.listdir(self.fastq_directory):
            if fastq_file.endswith(".fastq.gz"):
                with gzip.open(self.fastq_directory + "/" + fastq_file, mode='rt') as fastq_file_obj:

                    #Dictionary for saving different fields before writing to model
                    core_field_dict = {}

                    # Function defined above
                    try:
                        parse_illumina_fastq(fastq_file_obj, core_field_dict)
                        numof_files_parsed += 1
                    except:
                        raise

                    #Save filename
                    core_field_dict['filename'] = fastq_file

                    #Save project_id from chosen work order object
                    core_field_dict['project_id'] = self.project_id

                    top_seq_identifier = ""
                    seq_id_list = []

                    ## Filename operations

                    # Divide filename into components
                    illumina_split = fastq_file.split("_")

                    # Extract read direction
                    if illumina_split[-2] == 'R1' or illumina_split[-2] == 'R2':
                        core_field_dict['read'] = illumina_split[-2]
                    # Extract lane number
                    if illumina_split[-3].startswith('L'):
                        core_field_dict['lane'] = illumina_split[-3]
                    # Extract sample ID
                    core_field_dict['sample_id'] = _split[-1]
                    # Extract pool ID
                    core_field_dict['pool_id'] = _split[-2]

                    ## Create and populate a model object for this fastq file
                    data = CoreData.objects.create(**core_field_dict)

        return numof_files_parsed

class DataComparison(object):
    """Compares all core data objects to customer sample objects

    Comparison based on barcodes - returns list of barcodes with no complement.
    Uses fuzzy string matching in case first read is N.
    """

    def __init__(self, wo_to_compare):
        self.wo_to_compare = wo_to_compare

    def compare_data(self):

        core_data_objects = CoreData.objects.filter(project_id=self.wo_to_compare)
        customer_objects = ComponentInformation.objects.filter(project_id=self.wo_to_compare)

        core_index_list = []
        customer_index_list = []

        for obj in core_data_objects:
            core_index_list.append((str(obj.i7_index_sequence), str(obj.i5_index_sequence)))

        for obj in customer_objects:
            customer_index_list.append((str(obj.i7_index_sequence), str(obj.i5_index_sequence)))

        self.match_num = 0
        match_list = []
        self.no_match_core = []
        self.no_match_cust = []

        for cust_indexes in customer_index_list:

            found_match = False

            for core_indexes in core_index_list:
                # Matches first and second indexes with an 80% ratio
                if fuzz.ratio(cust_indexes[0], core_indexes[0]) > 80 and \
                   fuzz.ratio(cust_indexes[1], core_indexes[1]) > 80:
                        self.match_num += 1

                        # Add the customer and core indexes into the match_list
                        if cust_indexes not in match_list and \
                           core_indexes not in match_list:
                            match_list.append(cust_indexes)
                            match_list.append(core_indexes)
                        found_match = True
                        break

            if not found_match:
                # Aggregates indexes with no matches to anything
                self.no_match_cust.append(cust_indexes)

        for core_indexes in core_index_list:
            if core_indexes not in match_list:
                self.no_match_core.append(core_indexes)

        self.comparison_output = "\n".join([
            "Cust Indexes: %s" % len(customer_index_list),
            "Core Indexes: %s" % len(core_index_list),
            "Matches: %s" % self.match_num,
            "Customer Indexes with no matches: %s" % self.no_match_cust,
            "Core Indexes with no matches: %s" % self.no_match_core
        ])

def error_logger(error, project_id):
    """Creates database entries for failed / erroneous runs"""
    ExecutionStats.objects.create(
        project_id = project_id,
        exec_date = datetime.datetime.now(),
        exec_status = 'FAIL',
        fail_reason = error
    )

def cell_value(column, row, worksheet):
    return worksheet.cell(column=column, row=row).value
